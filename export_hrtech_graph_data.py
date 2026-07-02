from __future__ import annotations

import json
import math
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCES = [
    {
        "path": ROOT.parent / "outputs" / "hrtech_startups_master" / "hrtech_startups_master_taxonomy.xlsx",
        "google_sheet_url": "https://docs.google.com/spreadsheets/d/1vgIrAqAjl-QDRl90Uu1EDf8heXk1HKVQA0WPokatY9M",
    },
]
TARGET = ROOT / "graph-data.js"
TAXONOMY = ROOT.parent / "outputs" / "hrtech_startups_master" / "domain_taxonomy.json"

TAXONOMY_DATA = json.loads(TAXONOMY.read_text(encoding="utf-8"))
CATEGORY_META = {domain["name"]: domain for domain in TAXONOMY_DATA["domains"]}


def scalar(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def rows(ws):
    raw_rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell or "").strip() for cell in raw_rows[0]]
    for raw in raw_rows[1:]:
        item = {headers[i]: scalar(raw[i] if i < len(raw) else "") for i in range(len(headers))}
        if any(item.values()):
            yield item


def as_number(value, default=0):
    if value in ("", None):
        return default
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return default


def parse_category_weights(value, fallback):
    text = str(value or "").strip()
    if text:
        try:
            parsed = json.loads(text)
            result = [
                {
                    "category": str(item.get("domain") or item.get("category")),
                    "weight": as_number(item.get("weight"), 0),
                }
                for item in parsed
                if item.get("domain") or item.get("category")
            ]
            if result:
                return result
        except (json.JSONDecodeError, TypeError, AttributeError):
            pass
    return [{"category": fallback, "weight": 1}]


def startup_from_row(row):
    amount_m = as_number(row.get("amount_m"), 0)
    is_mna = str(row.get("mna_flag", "")).strip().lower() == "yes"
    tags = [str(row.get(f"tag_{i}", "")).strip() for i in range(1, 6)]
    category = str(row["category"])
    category_weights = parse_category_weights(row.get("category_weights"), category)
    return {
        "id": str(row["startup_id"]),
        "name": str(row["startup_name"]),
        "nodeModel": str(row.get("node_model", "Deal") or "Deal"),
        "companyKey": str(row.get("company_key", "")),
        "sameCompanyDealCount": int(as_number(row.get("same_company_deal_count"), 1)),
        "website": str(row["website"]),
        "summary": str(row["product_summary"]),
        "category": category,
        "legacyCategory": str(row.get("legacy_category", "")),
        "group": str(row["category_group"]),
        "domains": [item["category"] for item in category_weights],
        "categoryWeights": category_weights,
        "tags": [tag for tag in tags if tag],
        "tagsText": str(row["tags_5"]),
        "year": int(as_number(row["investment_year"], 0)),
        "quarter": str(row["investment_quarter"]),
        "month": str(row["investment_month"]),
        "dealType": str(row["deal_type"]),
        "investmentOriginal": str(row["investment_amount_original"]),
        "amountCurrency": str(row["amount_currency"]),
        "amountM": amount_m,
        "amountBucket": str(row["amount_bucket"]),
        "mna": is_mna,
        "acquirer": str(row["acquirer"]),
        "dealSourceUrl": str(row["deal_source_url"]),
        "semanticSourceUrl": str(row["semantic_source_url"]),
        "confidence": str(row["confidence"]),
    }


def main():
    startups = []
    categories = {}
    rules_by_id = {}

    for source in SOURCES:
        wb = load_workbook(source["path"], read_only=True, data_only=True)

        for row in rows(wb["Startups"]):
            startups.append(startup_from_row(row))

        for row in rows(wb["Categories"]):
            category = str(row["category"])
            categories.setdefault(
                category,
                {
                    "group": str(row["category_group"]),
                    "description": str(row["description"]),
                    "count": 0,
                    "color": CATEGORY_META.get(category, {}).get("color", "#9CA3AF"),
                    "anchor": {
                        "x": CATEGORY_META.get(category, {}).get("x", 0.5),
                        "y": CATEGORY_META.get(category, {}).get("y", 0.5),
                    },
                },
            )

        for row in rows(wb["Rules"]):
            rule_id = str(row["rule_id"])
            rules_by_id.setdefault(
                rule_id,
                {
                    "id": rule_id,
                    "name": str(row["rule_name"]),
                    "logic": str(row["logic"]),
                    "action": str(row["edge_weight_or_action"]),
                },
            )

    for category in categories:
        categories[category]["count"] = sum(
            1
            for item in startups
            if item["category"] == category
        )
    categories = {name: meta for name, meta in categories.items() if meta["count"] > 0}

    tag_df = Counter()
    for item in startups:
        tag_df.update(set(item["tags"]))

    corpus_size = len(startups) or 1

    def tag_specificity(tag):
        return math.log((corpus_size + 1) / (tag_df[tag] + 1)) / math.log(corpus_size + 1)

    relations = []
    for i, source in enumerate(startups):
        source_tags = set(source["tags"])
        for target in startups[i + 1 :]:
            if source.get("companyKey") and target.get("companyKey") and source["companyKey"] == target["companyKey"]:
                continue
            shared_tags = sorted(source_tags.intersection(target["tags"]))
            if not shared_tags:
                continue
            same_category = source["category"] == target["category"]
            relation_weight = sum(tag_specificity(tag) for tag in shared_tags)
            strongest_shared_tag = max(tag_specificity(tag) for tag in shared_tags)
            if relation_weight >= 1.5 or (len(shared_tags) >= 3 and relation_weight >= 1.15 and strongest_shared_tag >= 0.25):
                relations.append(
                    {
                        "source": source["id"],
                        "target": target["id"],
                        "type": "idf_tags_same_domain" if same_category else "idf_tags",
                        "sharedTags": shared_tags,
                        "weight": round(relation_weight, 3),
                        "strength": max(1, min(5, round(1 + min(1, relation_weight / 2.4) * 4))),
                        "ruleIds": "R3; R9",
                    }
                )

    relations.sort(
        key=lambda item: (
            -item["weight"],
            -item["strength"],
            -len(item["sharedTags"]),
            item["source"],
            item["target"],
        )
    )
    relations = strongest_relations_per_node(relations, max_per_node=5)

    quarter_counts = {}
    for item in startups:
        quarter_counts[item["quarter"]] = quarter_counts.get(item["quarter"], 0) + 1

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceWorkbooks": [
            {
                "path": str(source["path"]),
                "googleSheetUrl": source["google_sheet_url"],
            }
            for source in SOURCES
        ],
        "quarterCounts": quarter_counts,
        "startups": startups,
        "relations": relations,
        "categories": categories,
        "rules": [rules_by_id[key] for key in sorted(rules_by_id)],
        "amountBuckets": ["Under 5M", "5-9.9M", "10-19.9M", "20-49.9M", "50M+", "M&A / undisclosed"],
        "years": sorted({item["year"] for item in startups if item["year"]}),
    }
    TARGET.write_text(
        "window.HRTECH_GRAPH_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(
        f"Generated {TARGET}: {len(startups)} startups, "
        f"{len(relations)} relations, {len(categories)} categories, quarters={quarter_counts}"
    )


def strongest_relations_per_node(relations, max_per_node=8):
    counts = {}
    kept = []
    for relation in relations:
        source_count = counts.get(relation["source"], 0)
        target_count = counts.get(relation["target"], 0)
        if source_count >= max_per_node and target_count >= max_per_node:
            continue
        kept.append(relation)
        counts[relation["source"]] = source_count + 1
        counts[relation["target"]] = target_count + 1
    return kept


if __name__ == "__main__":
    main()
